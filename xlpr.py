import argparse
import xlrd
import os
from copy import copy

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule

parser = argparse.ArgumentParser(description='Automate creation of scoring Excel worksheets.')
subparsers = parser.add_subparsers(dest='subcommand', title='subcommands', description='valid subcommands', required=True)

manual = subparsers.add_parser('manual')
manual.add_argument('name', help='What to call the file')
manual.add_argument('num_questions', type=int, help='How many question columns to generate')
manual.add_argument('num_participants', type=int, help='How many participant rows to generate')
manual.add_argument('--range-high', dest='range_high', type=int, help='If there is a range for all questions, what is the top of that range for conditional formatting?')

auto = subparsers.add_parser('auto')
auto.add_argument('input', help='Input excel file that specifies what questionairres to create')
auto.add_argument('num_participants', type=int, help='How many participant rows to generate')
auto.add_argument('output_path', help='Where to create the files')

addcols = subparsers.add_parser('addcols')
addcols.add_argument('input', help='Questionairre file to alter')
addcols.add_argument('num_columns', type=int, help='How many columns to add')

rebuild = subparsers.add_parser('rebuild')
rebuild.add_argument('input', help='Questionairre file to alter')

rebuild = subparsers.add_parser('dr')
rebuild.add_argument('input', help='Day reconstruction file to alter')

check = subparsers.add_parser('check')
check.add_argument('input', help='Questionairre file to check')

args = parser.parse_args()

small_font = Font(size = "9")
border_right = Border(right=Side(style='double', color='0000ff'))
border_bottom = Border(bottom=Side(style='double', color='0000ff'))


def do_borders(ws, cols, rows):
    for col in range(6, cols):
        ws.cell(column=col, row=3).border = border_bottom
    for row in range(4, rows):
        ws.cell(column=5, row=row).border = border_right

def cf_highlight_good_row(ws, first_cell, last_cell, start_row, end_row):
    # Conditional formatting to highlight rows with no mismatches
    greenFill = PatternFill(start_color='66EE66', end_color='66EE66', fill_type='solid')
    dxf = DifferentialStyle(fill=greenFill)
    rule = Rule(type="expression", formula = ["COUNT(SEARCH(\"|\",${0}:${1}))<1".format(start_row, end_row)], dxf=dxf, stopIfTrue=True)
    ws.conditional_formatting.add('{0}:{1}'.format(first_cell, last_cell), rule)

def cf_mismatches(ws, first_cell, last_cell):
    # Conditional formatting to highlight mismatches
    redFill = PatternFill(start_color='EE6666', end_color='EE6666', fill_type='solid')
    dxf = DifferentialStyle(fill=redFill)
    rule = Rule(type="expression", formula = ["ISNUMBER(SEARCH(\"|\", {0}))".format(first_cell)], dxf=dxf, stopIfTrue=True)
    ws.conditional_formatting.add('{0}:{1}'.format(first_cell, last_cell), rule)

def cf_blanks(ws, first_cell, last_cell):
    # Conditional formatting to highlight blanks as light gray
    blankFill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
    dxf = DifferentialStyle(fill=blankFill)
    blank = Rule(type="expression", formula = ["ISBLANK({0})".format(first_cell)], dxf=dxf, stopIfTrue=True)
    ws.conditional_formatting.add('{0}:{1}'.format(first_cell, last_cell), blank)

def fill_sheet(ws, num_questions, num_participants, range_high, range_low=1, copyRowsFromSheet1=False):
    # Add headers
    ws['A1'] = 'Sub ID'
    ws['B1'] = 'Session Date'
    ws['C1'] = 'Visit'
    ws['D1'] = 'Date Entered'
    ws['E1'] = 'Entered By'
    for col in range(1, 6 + num_questions):
        ws.column_dimensions[get_column_letter(col)].width = "15"
    for col in range(6, 6 + num_questions):
        question_metadata1 = ws.cell(column=col,row=2)
        question_metadata1.font = small_font
        question_metadata2 = ws.cell(column=col,row=3)
        question_metadata2.font = small_font
        if copyRowsFromSheet1:
            ws.cell(column=col, row=1, value='=Entry1!{0}{1}'.format(get_column_letter(col), 1))
            question_metadata1.value='=Entry1!{0}{1}'.format(get_column_letter(col), 2)
            question_metadata2.value='=Entry1!{0}{1}'.format(get_column_letter(col), 3)
        else:
            ws.cell(column=col, row=1, value="Q{0}".format(col-5))

    ws.cell(column=6 + num_questions, row=1, value="Notes")

    # Add subject numbers
    for row in range(4, 4 + num_participants):
        ws.cell(column=1, row=row, value=(row-3+1000))

    # This freezes everything "above and to the left" of the given cell
    ws.freeze_panes = "F4"

    do_borders(ws, 6 + num_questions, 4 + num_participants)
    
    first_cell = ws.cell(column=6, row=3)
    last_cell = ws.cell(column=5 + num_questions, row=2 + num_participants)

    # Conditional formatting to do blanks as light gray
    cf_blanks(ws, 'F4', last_cell.coordinate)
                              
    # Conditional formatting to highlight numbers out of range
    if range_high:
        yellowFill = PatternFill(start_color='EEEE66', end_color='EEEE66', fill_type='solid')
        rule = CellIsRule(operator="notBetween", formula=[str(range_low),str(range_high)], fill=yellowFill)
        ws.conditional_formatting.add('F4:{0}'.format(last_cell.coordinate), rule)

    return ws

def compare_cell(cell, is_date=False, name1='Entry1', name2='Entry2'):
    if is_date:
        cell.value = "=IF({1}!{0}={2}!{0},IF(ISBLANK({1}!{0}),\"\",TEXT({1}!{0},\"mm/dd/yyyy\")),CONCATENATE(TEXT({1}!{0},\"mm/dd/yyyy\"),\" | \",TEXT({2}!{0},\"mm/dd/yyyy\")))".format(cell.coordinate, "'" + name1 + "'", "'" + name2 + "'")
    else:
        cell.value = "=IF({1}!{0}={2}!{0},IF(ISBLANK({1}!{0}),\"\",{2}!{0}),CONCATENATE({1}!{0},\" | \",{2}!{0}))".format(cell.coordinate, "'" + name1 + "'", "'" + name2 + "'")


def compare_sheet(ws, num_questions, num_participants,
        skip_columns=6, name1='Entry1', name2='Entry2',
        is_day_reconstruction=False, header_sheet=None):
    # Add headers
    ws['A1'] = 'Sub ID'
    if is_day_reconstruction:
        ws['B1'] = 'Rater 1'
        ws['C1'] = 'Rater 2'
        ws['D1'] = 'Day'
        # Cell for test comparison
        ws['B2'] = 'Test|Comparison'
    else:
        ws['B1'] = 'Session Date'
        ws['C1'] = 'Visit'
        ws['D1'] = 'Rater 1'
        ws['E1'] = 'Rater 2'
        # Cell for test comparison
        ws['B3'] = 'Test|Comparison'

    formula1 = '=IF(ISBLANK(\'' + name1 + '\'!{0}{1}),"",\'' + name1 + '\'!{0}{1})'
    formula2 = '=IF(ISBLANK(\'' + name2 +  '\'!{0}{1}),"",\'' + name2 + '\'!{0}{1})'

    for col in range(1, skip_columns + num_questions):
        ws.column_dimensions[get_column_letter(col)].width = "15"

    # Headings are taken care of by copy_headings

    # Look up if any of the headers should be dates, for formatting formulas
    date_columns = []
    if header_sheet:
        for col in range(1, skip_columns + num_questions):
            header = header_sheet.cell(column=col,row=2).value
            if header:
                if "date" in header: 
                    date_columns.append(col)

    for row in range(4, 4 + num_participants):
        subject = ws.cell(column=1,row=row)
        compare_cell(subject, name1=name1, name2=name2)
        if is_day_reconstruction:
            rater1 = ws.cell(column=2,row=row)
            rater1.value = formula1.format("C", row)
            rater2 = ws.cell(column=3,row=row)
            rater2.value = formula2.format("C", row)
        else:
            session_date = ws.cell(column=2,row=row)
            compare_cell(session_date, is_date=True, name1=name1, name2=name2)
            visit = ws.cell(column=3,row=row)
            compare_cell(visit, name1=name1, name2=name2)
            rater1 = ws.cell(column=4,row=row)
            rater1.value = formula1.format("E", row)
            rater2 = ws.cell(column=5,row=row)
            rater2.value = formula2.format("E", row)
        for col in range(skip_columns, skip_columns + num_questions):
            cell = ws.cell(column=col,row=row)
            is_date = col in date_columns
            compare_cell(cell, is_date=is_date, name1=name1, name2=name2)

    last_cell = ws.cell(column=4 + num_questions, row=3 + num_participants)

    cf_highlight_good_row(ws, 'A4',
            ws.cell(column=1, row=3 + num_participants).coordinate,
            ws.cell(column=5, row=4).coordinate,
            ws.cell(column=3 + num_questions, row=4).coordinate)
    cf_mismatches(ws, 'B3', last_cell.coordinate)

    # Protect the sheet so nobody can edit it
    ws.protection.sheet = True

    return ws


def generate_workbook(name, num_questions, num_participants, range_high, range_low=1):
    wb = Workbook()

    wb.remove(wb.active)
    entry1 = fill_sheet(wb.create_sheet("Entry1"), num_questions, num_participants, range_high, range_low)
    entry2 = fill_sheet(wb.create_sheet("Entry2"), num_questions, num_participants, range_high, range_low, copyRowsFromSheet1=True)
    compare = compare_sheet(wb.create_sheet("Final_Comparison"), num_questions, num_participants)
    copy_headings(entry1, compare)

    if not ".xlsx" in name:
        name = name + ".xlsx"
    wb.save(name)


def generate_manual():
    generate_workbook(args.name, args.num_questions, args.num_participants, args.range_high)


def generate_automatic():
    num_participants = args.num_participants

    # load file
    workbook = xlrd.open_workbook(args.input)
    sheet = workbook.sheet_by_index(0)

    for row in range(1,sheet.nrows):
        num_questions = sheet.cell_value(row,4)
        if num_questions == '': continue
        try:
            num_questions = int(num_questions)
        except ValueError:
            print(f'Ignoring row {row}, "{num_questions}" is not a question number I understand')
            continue
        if num_questions > 0:
            name = sheet.cell_value(row,1)
            consistent_scale = sheet.cell_value(row,5)
            if consistent_scale == 'y':
                low = sheet.cell_value(row,6)
                high = sheet.cell_value(row,7)
            else:
                low = None
                high = None

            if low == "":
                low = None
            if high == "":
                high = None

            generate_workbook(os.path.join(args.output_path, name), num_questions, num_participants, high, low)


def add_columns_to_existing_workbook():
    wb = load_workbook(args.input)
    sheet1 = wb.worksheets[0]
    sheet2 = wb.worksheets[1]
    sheet3 = wb.worksheets[2]

    col_extent = sheet1.max_column
    row_extent = sheet1.max_row

    sheet1.insert_cols(col_extent, args.num_columns)
    sheet2.insert_cols(col_extent, args.num_columns)
    sheet3.insert_cols(col_extent, args.num_columns)

    for col in range(col_extent, col_extent + args.num_columns):
        sheet1.cell(column=col, row=1, value="Q{0}".format(col-5))
        question_metadata1 = sheet1.cell(column=col,row=2)
        question_metadata1.font = small_font
        question_metadata2 = sheet1.cell(column=col,row=3)
        question_metadata2.font = small_font

    first_cell = sheet1.cell(column=col_extent, row=4)
    last_cell = sheet1.cell(column=col_extent + args.num_columns, row=row_extent)
    cf_blanks(sheet1, first_cell.coordinate, last_cell.coordinate)

    def copy_sheet(sheet, compare):
        if compare:
            cf_mismatches(sheet, first_cell.coordinate, last_cell.coordinate)
        else:
            cf_blanks(sheet, first_cell.coordinate, last_cell.coordinate)
        # Conditional formatting
        for row in range(1, row_extent):
            for col in range(col_extent, col_extent + args.num_columns):
                cell = sheet.cell(column=col,row=row)
                if row <= 3:
                    cell.value='=Entry1!{0}{1}'.format(get_column_letter(col), row)
                if row == 2 or row == 3:
                    cell.font = small_font
                elif row > 3 and compare:
                    compare_cell(cell)

    copy_sheet(sheet2, False)
    copy_sheet(sheet3, True)

    def widen_sheet(sheet):
        for col in range(col_extent, col_extent + args.num_columns):
            sheet.column_dimensions[get_column_letter(col)].width = "15"

    widen_sheet(sheet1)
    widen_sheet(sheet2)
    widen_sheet(sheet3)

    do_borders(sheet1, col_extent + args.num_columns, row_extent)
    do_borders(sheet2, col_extent + args.num_columns, row_extent)
    do_borders(sheet3, col_extent + args.num_columns, row_extent)

    wb.save(args.input)


def copy_headings(sheet1, sheet2, start_row=4, start_col=5):
    col_extent = sheet1.max_column
    row_extent = sheet1.max_row

    for row in range(1, start_row):
        # sheet2.row_dimensions[1].height = sheet1.row_dimensions[1].height
        sheet2.row_dimensions = sheet1.row_dimensions
        for col in range(start_col, col_extent):
            source = sheet1.cell(column=col, row=row)
            target = sheet2.cell(column=col, row=row)
            target.value = source.value
            if source.has_style:
                target._style = copy(source._style)


def rebuild_existing_workbook():
    wb = load_workbook(args.input)
    first = wb.worksheets[0]
    second = wb.worksheets[1]
    compare = wb.worksheets[2]

    col_extent = first.max_column
    row_extent = first.max_row

    num_questions = col_extent - 4
    num_participants = row_extent - 3

    compare.protection.sheet = False
    wb.remove(compare)

    copy_headings(first, second)
    print(f"Building comparison sheet with {num_questions} questions and {num_participants} participants")
    compare_sheet(wb.create_sheet("Final_Comparison"), num_questions, num_participants)
    copy_headings(first, wb.worksheets[2])
    wb.save(args.input)


def check_workbook():
    try:
        wb = load_workbook(args.input)
        first = wb.worksheets[0]
        second = wb.worksheets[1]

        columns1 = first.max_column
        rows1 = first.max_row

        columns2 = second.max_column
        rows2 = second.max_row

        if columns1 == columns2 and rows1 == rows2:
            print(f"Found {columns1} columns and {rows1} rows in {args.input}")
        else:
            print(f"ERROR, {columns1} columns and {rows1} rows on first sheet and {columns2} columns and {rows2} rows on second sheet in {args.input}")

        mismatches = []

        def checker(col,row):
            r = second.cell(column=col,row=row)
            expected = '=Entry1!{0}{1}'.format(get_column_letter(col), row)
            if r.value != expected:
                mismatches.append(r.coordinate)

        # Check if sheet two first three columns are correct
        # Skip notes column
        for col in range(6, columns2 - 1):
            checker(col, 1)
            checker(col, 2)
            checker(col, 3)

        if mismatches:
            print(f"WARNING: Formula mismatches! {mismatches}")

    except Exception as e:
        print(f"Error in {args.input}: {e}")


def day_reconstructions():
    wb = load_workbook(args.input)

    first = wb.worksheets[wb.sheetnames.index("R1_day")]

    num_questions = 21
    num_rows = first.max_row + 100

    compare = wb.worksheets[wb.sheetnames.index("Days_comparison")]
    compare.protection.sheet = False
    wb.remove(compare)

    print(f"Building days comparison sheet with {num_questions} questions and {num_rows} rows")
    days_comparison = wb.create_sheet("Days_comparison", index=2)
    ws = compare_sheet(
            days_comparison,
            num_questions, num_rows,
            skip_columns=4, is_day_reconstruction=True,
            header_sheet=first,
            name1='R1_day', name2='R2_day'
            )

    copy_headings(first, days_comparison, start_row=4, start_col=5)

    # Now the episodes pair
    compare = wb.worksheets[wb.sheetnames.index("Episodes_comparison")]
    compare.protection.sheet = False
    wb.remove(compare)

    first_episodes = wb.worksheets[wb.sheetnames.index("R1_episodes")]

    num_questions = 48
    num_rows = first_episodes.max_row + 500
    # Sometimes max row is waaaay too high
    if num_rows > 1000:
        num_rows = 1000

    print(f"Building episodes comparison sheet with {num_questions} questions and {num_rows} rows")
    episodes_comparison = wb.create_sheet("Episodes_comparison", index=5)
    ws = compare_sheet(
            episodes_comparison,
            num_questions, num_rows,
            skip_columns=4, is_day_reconstruction=True,
            header_sheet=first_episodes,
            name1='R1_episodes', name2='R2_episodes'
            )

    copy_headings(first_episodes, episodes_comparison, start_row=4, start_col=5)
    wb.save(args.input)


if args.subcommand == 'auto':
    generate_automatic()
elif args.subcommand == 'addcols':
    add_columns_to_existing_workbook()
elif args.subcommand == 'rebuild':
    rebuild_existing_workbook()
elif args.subcommand == 'check':
    check_workbook()
elif args.subcommand == 'dr':
    day_reconstructions()
else:
    generate_manual()
